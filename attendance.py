'''
1. 将 xls 转换成 xlsx 格式
2. 读取“刷卡记录”sheet
3. 放入数据库所有人的记录(记录格式：[姓名，日期，上班时间，下班时间，工号，部门])
'''
import re,json,sqlite3
from sqlite3.dbapi2 import Cursor
import win32com.client as win32
from openpyxl import load_workbook

#将xls转换成xlsx
fname = r"C:\Users\zcy\Desktop\pythonTask\09月汇总表.xls"#路径前加r保持原意不转义
excel = win32.gencache.EnsureDispatch('Excel.Application')
wbt = excel.Workbooks.Open(fname)

wbt.SaveAs(fname+"x", FileFormat = 51)    #FileFormat = 51 is for .xlsx extension
wbt.Close()                               #FileFormat = 56 is for .xls extension
excel.Application.Quit()

wb = load_workbook(filename="09月汇总表.xlsx")
ws = wb['刷卡记录']

#读取json信息
file = open('./南京.json','r',encoding='utf8')
json_data = json.load(file)
map={'275':321,'276':322,'331':342,'332':341,'347':346,'348':347}

#数据库创建连接
conn=sqlite3.connect('attendance_sheet.db')
c=conn.cursor()
c.execute('create table record(name text(20),date text(20),sign_in text(20),sign_out text(20),special text(20),jobnumber text(10),department text(20))')
conn.commit()
conn.close()

# [姓名，日期，上班时间，下班时间，工号，部门]
nd = 3  
for m in range(1, 39):  # 员工
    nd += 2
    for n in range(1, 31):  # 日期
        job_number = ws.cell(nd, 3).value
        if(job_number in map):
            temp=map['%s'%job_number]
            job_number=temp
            num=json_data['%s'%temp]
        else:
            num=json_data['%s'%job_number]
        name=num[0]
        department=num[1]
        #日期处理
        date_cell=ws.cell(3,29).value
        date=date_cell[0:8]
        date=date.replace("/","-")
        if(n//10<1):
            day='0%d'%n
        else:
            day=n
        date_cell_value=f'{date}{day}'
        
        clock=ws.cell(nd+1, n)
        val=clock.value
        if(val):
            val=val
        else:
            val=" \n"
        pattern=re.compile(r'[^\n]+')
        result=pattern.findall(val)
        length=len(result)

        if(length==1):
            clock_in_value = result[0]
            clock_out_value = result[0]
        elif(length==2):
            clock_in_value = result[0]
            clock_out_value = result[1]
        elif(result[1]<"12:00"):
            clock_in_value = result[1]
            clock_out_value = result[2]
        else:
            clock_in_value = result[0]
            clock_out_value = result[2]

        conn1=sqlite3.connect('attendance_sheet.db')
        c1=conn1.cursor()
        
        special=''
        if(clock_in_value==" " and clock_out_value==" "):
            pass
        else:
            #print(f'[姓名:{name}, 日期:{date_cell_value}, 部门:{department}, 工号:{job_number}, 上班时间:{clock_in_value}, 下班时间:{clock_out_value}]')
            c1.execute(f"INSERT INTO record (name,date,department,jobnumber,sign_in,sign_out,special) VALUES ('{name}', '{date_cell_value}','{department}', '{job_number}', '{clock_in_value}', '{clock_out_value}','{special}' )")
            conn1.commit()
            conn1.close()


""" #查询信息
conn2=sqlite3.connect('attendance_sheet.db')
search_name=input("请输入想要查询考勤信息的名字:")
cursor=conn2.execute('select * from record where name="%s"'%search_name)
for row in cursor:
    print(row)
conn2.close() """


